import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    def __init__(self, output_dir="Test Results/Excel"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Styles
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.pass_font = Font(name="Calibri", size=10, color="006100")
        
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.fail_font = Font(name="Calibri", size=10, color="9C0006")
        
        self.skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.skip_font = Font(name="Calibri", size=10, color="9C6500")
        
        self.bold_font = Font(name="Calibri", size=11, bold=True)
        self.regular_font = Font(name="Calibri", size=10)
        self.thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

    def _style_header(self, ws, columns):
        ws.append(columns)
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    def generate_automation_test_report(self, test_results, file_path=None):
        if not file_path:
            file_path = os.path.join(self.output_dir, "Automation_Test_Report.xlsx")
            
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet
        
        # Filter test result groups
        passed_tests = [t for t in test_results if t.get("status") == "PASSED"]
        failed_tests = [t for t in test_results if t.get("status") == "FAILED"]
        skipped_tests = [t for t in test_results if t.get("status") == "SKIPPED"]
        
        # Sheet 1: Executed Test Cases
        ws1 = wb.create_sheet(title="Executed Test Cases")
        cols1 = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)", "Failure Reason"]
        self._style_header(ws1, cols1)
        for t in test_results:
            row_idx = ws1.max_row + 1
            ws1.append([
                t.get("test_id"), t.get("module"), t.get("test_name"),
                t.get("priority"), t.get("status"), t.get("execution_time", 0.05),
                t.get("failure_reason", "N/A")
            ])
            status_cell = ws1.cell(row=row_idx, column=5)
            if t.get("status") == "PASSED":
                status_cell.fill, status_cell.font = self.pass_fill, self.pass_font
            elif t.get("status") == "FAILED":
                status_cell.fill, status_cell.font = self.fail_fill, self.fail_font
            else:
                status_cell.fill, status_cell.font = self.skip_fill, self.skip_font
        self._auto_fit_columns(ws1)
        
        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title="Passed Tests")
        self._style_header(ws2, cols1[:6])
        for t in passed_tests:
            row_idx = ws2.max_row + 1
            ws2.append([t.get("test_id"), t.get("module"), t.get("test_name"), t.get("priority"), t.get("status"), t.get("execution_time", 0.05)])
            ws2.cell(row=row_idx, column=5).fill = self.pass_fill
            ws2.cell(row=row_idx, column=5).font = self.pass_font
        self._auto_fit_columns(ws2)
        
        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title="Failed Tests")
        self._style_header(ws3, cols1)
        for t in failed_tests:
            row_idx = ws3.max_row + 1
            ws3.append([t.get("test_id"), t.get("module"), t.get("test_name"), t.get("priority"), t.get("status"), t.get("execution_time", 0.05), t.get("failure_reason", "N/A")])
            ws3.cell(row=row_idx, column=5).fill = self.fail_fill
            ws3.cell(row=row_idx, column=5).font = self.fail_font
        self._auto_fit_columns(ws3)
        
        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title="Skipped Tests")
        self._style_header(ws4, cols1)
        for t in skipped_tests:
            row_idx = ws4.max_row + 1
            ws4.append([t.get("test_id"), t.get("module"), t.get("test_name"), t.get("priority"), t.get("status"), t.get("execution_time", 0.0), t.get("failure_reason", "Feature Disabled / Skipped")])
            ws4.cell(row=row_idx, column=5).fill = self.skip_fill
            ws4.cell(row=row_idx, column=5).font = self.skip_font
        self._auto_fit_columns(ws4)
        
        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        self._style_header(ws5, ["Metric Name", "Value"])
        total = len(test_results)
        pass_count = len(passed_tests)
        fail_count = len(failed_tests)
        skip_count = len(skipped_tests)
        pass_rate = round((pass_count / total * 100), 2) if total > 0 else 0.0
        
        metrics = [
            ["Total Executed Test Cases", total],
            ["Passed Test Cases", pass_count],
            ["Failed Test Cases", fail_count],
            ["Skipped Test Cases", skip_count],
            ["Pass Percentage", f"{pass_rate}%"],
            ["Framework Target", "Enterprise Mobile & Web & Security"],
            ["Execution Mode", "Headless / Automated Pipeline"]
        ]
        for m in metrics:
            ws5.append(m)
        self._auto_fit_columns(ws5)
        
        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        self._style_header(ws6, ["Defect ID", "Test Case ID", "Module", "Severity", "Description", "Status"])
        for idx, f_test in enumerate(failed_tests, 1):
            ws6.append([
                f"DEF-00{idx}", f_test.get("test_id"), f_test.get("module"),
                f_test.get("priority", "High"), f_test.get("failure_reason", "Assertion Failure"), "Open"
            ])
        self._auto_fit_columns(ws6)

        # Sheet 7: Pass Rate Summary
        ws7 = wb.create_sheet(title="Pass Rate Summary")
        self._style_header(ws7, ["Module", "Total Tests", "Passed", "Failed", "Pass Rate (%)"])
        modules = sorted(list(set(t.get("module", "General") for t in test_results)))
        for mod in modules:
            mod_tests = [t for t in test_results if t.get("module") == mod]
            mod_passed = len([t for t in mod_tests if t.get("status") == "PASSED"])
            mod_failed = len([t for t in mod_tests if t.get("status") == "FAILED"])
            mod_rate = round((mod_passed / len(mod_tests) * 100), 2) if mod_tests else 0.0
            ws7.append([mod, len(mod_tests), mod_passed, mod_failed, f"{mod_rate}%"])
        self._auto_fit_columns(ws7)
        
        wb.save(file_path)
        
        # Also save Passed_Test_Cases.xlsx, Failed_Test_Cases.xlsx, Execution_Summary.xlsx
        wb_pass = openpyxl.Workbook()
        ws_p = wb_pass.active
        ws_p.title = "Passed Tests"
        self._style_header(ws_p, cols1[:6])
        for t in passed_tests:
            ws_p.append([t.get("test_id"), t.get("module"), t.get("test_name"), t.get("priority"), t.get("status"), t.get("execution_time", 0.05)])
        self._auto_fit_columns(ws_p)
        wb_pass.save(os.path.join(self.output_dir, "Passed_Test_Cases.xlsx"))
        
        wb_fail = openpyxl.Workbook()
        ws_f = wb_fail.active
        ws_f.title = "Failed Tests"
        self._style_header(ws_f, cols1)
        for t in failed_tests:
            ws_f.append([t.get("test_id"), t.get("module"), t.get("test_name"), t.get("priority"), t.get("status"), t.get("execution_time", 0.05), t.get("failure_reason", "N/A")])
        self._auto_fit_columns(ws_f)
        wb_fail.save(os.path.join(self.output_dir, "Failed_Test_Cases.xlsx"))
        
        wb_sum = openpyxl.Workbook()
        ws_s = wb_sum.active
        ws_s.title = "Summary"
        self._style_header(ws_s, ["Metric", "Value"])
        for m in metrics:
            ws_s.append(m)
        self._auto_fit_columns(ws_s)
        wb_sum.save(os.path.join(self.output_dir, "Execution_Summary.xlsx"))

    def generate_vulnerability_excel_reports(self, vuln_dir, endpoint_list, security_findings, test_cases, dep_vulns, perf_results):
        os.makedirs(vuln_dir, exist_ok=True)
        
        # 1. endpoint-inventory.xlsx
        wb_ep = openpyxl.Workbook()
        ws_ep = wb_ep.active
        ws_ep.title = "Endpoint Inventory"
        cols_ep = ["Endpoint", "HTTP Method", "Authentication Required", "Expected Roles", "Controller / Source File"]
        self._style_header(ws_ep, cols_ep)
        for ep in endpoint_list:
            ws_ep.append([
                ep.get("endpoint"), ep.get("method"),
                "Yes" if ep.get("auth_required") else "No",
                ep.get("roles", "User/Admin"), ep.get("controller", "backend/main.py")
            ])
        self._auto_fit_columns(ws_ep)
        wb_ep.save(os.path.join(vuln_dir, "endpoint-inventory.xlsx"))
        
        # 2. findings.xlsx
        wb_fd = openpyxl.Workbook()
        ws_fd = wb_fd.active
        ws_fd.title = "Security Findings"
        cols_fd = ["Finding ID", "Severity", "Vulnerability Type", "CWE Mapping", "OWASP Mapping", "Endpoint / Component", "Description", "Status"]
        self._style_header(ws_fd, cols_fd)
        for f in security_findings:
            row_idx = ws_fd.max_row + 1
            ws_fd.append([
                f.get("id"), f.get("severity"), f.get("title"), f.get("cwe"),
                f.get("owasp"), f.get("endpoint"), f.get("description"), f.get("status", "Remediated/Documented")
            ])
            sev_cell = ws_fd.cell(row=row_idx, column=2)
            if f.get("severity") in ["Critical", "High"]:
                sev_cell.fill, sev_cell.font = self.fail_fill, self.fail_font
            elif f.get("severity") == "Medium":
                sev_cell.fill, sev_cell.font = self.skip_fill, self.skip_font
            else:
                sev_cell.fill, sev_cell.font = self.pass_fill, self.pass_font
        self._auto_fit_columns(ws_fd)
        wb_fd.save(os.path.join(vuln_dir, "findings.xlsx"))
        
        # 3. test-cases.xlsx (Master Security & Functional Test Cases Workbook)
        wb_tc = openpyxl.Workbook()
        ws_tc = wb_tc.active
        ws_tc.title = "Security & Functional Tests"
        cols_tc = ["Test Case ID", "Category", "Title", "Objective", "Preconditions", "Test Steps", "Expected Result", "Severity", "Status"]
        self._style_header(ws_tc, cols_tc)
        for tc in test_cases:
            row_idx = ws_tc.max_row + 1
            ws_tc.append([
                tc.get("test_id"), tc.get("category"), tc.get("title"), tc.get("objective"),
                tc.get("preconditions"), tc.get("steps"), tc.get("expected_result"), tc.get("severity"), tc.get("status")
            ])
            st_cell = ws_tc.cell(row=row_idx, column=9)
            if tc.get("status") == "PASSED":
                st_cell.fill, st_cell.font = self.pass_fill, self.pass_font
            elif tc.get("status") == "FAILED":
                st_cell.fill, st_cell.font = self.fail_fill, self.fail_font
            else:
                st_cell.fill, st_cell.font = self.skip_fill, self.skip_font
        self._auto_fit_columns(ws_tc)
        wb_tc.save(os.path.join(vuln_dir, "test-cases.xlsx"))
